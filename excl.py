import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import re

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed. Install it with: pip install openpyxl")

class ExcelLite:
    def __init__(self, rootexcellite2, file_path=None):
        self.rootexcellite2 = rootexcellite2
        self.rootexcellite2.title("Excel Lite")
        self.rootexcellite2.geometry("900x700")
        self.rootexcellite2.configure(bg="#c0c0c0")
        
        # Variables for tracking
        self.current_file = None
        self.workbook = None
        self.active_sheet = None
        self.cell_grid = []
        self.header_labels = []
        self.selected_cell = None
        self.selected_range = None
        self.formula_mode = False
        
        # Max rows and columns
        self.MAX_ROWS = 100
        self.MAX_COLS = 26  # A-Z
        
        # Column widths
        self.column_widths = [10] * (self.MAX_COLS + 1)
        
        # Cell data storage
        self.cell_data = {}
        
        # Clipboard
        self.clipboard_data = None
        
        # Context menu
        self.context_menu = None
        
        # Resize tracking
        self.resize_column = None
        self.resize_start_x = None
        self.resize_start_width = None
        self.resize_grip_width = 5  # pixels from edge to detect resize
        
        self.create_ui()
        
        # If file provided, open it
        if file_path and os.path.exists(file_path):
            self.open_excel_file(file_path)
    
    def _bind_mousewheel(self, event):
        """Bind mouse wheel when entering canvas"""
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)  # Linux
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)  # Linux

    def _unbind_mousewheel(self, event):
        """Unbind mouse wheel when leaving canvas"""
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling"""
        if event.num == 5 or event.delta < 0:
            # Scroll down
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            # Scroll up
            self.canvas.yview_scroll(-1, "units")
            
    def create_new_sheet(self):
        """Create a new sheet"""
        if not self.workbook:
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Get sheet name from user
        dialog = tk.Toplevel(self.rootexcellite2)
        dialog.title("New Sheet")
        dialog.geometry("300x120")
        dialog.transient(self.rootexcellite2)
        dialog.grab_set()
        
        tk.Label(dialog, text="Sheet name:", font=("MS Sans Serif", 9)).pack(pady=10)
        
        name_entry = tk.Entry(dialog, width=30, font=("MS Sans Serif", 9))
        name_entry.pack(pady=5)
        name_entry.insert(0, f"Sheet{len(self.workbook.sheetnames) + 1}")
        name_entry.select_range(0, tk.END)
        name_entry.focus_set()
        
        def create():
            sheet_name = name_entry.get().strip()
            if not sheet_name:
                messagebox.showwarning("Warning", "Sheet name cannot be empty")
                return
            
            if sheet_name in self.workbook.sheetnames:
                messagebox.showwarning("Warning", "Sheet name already exists")
                return
            
            # Create new sheet
            self.workbook.create_sheet(sheet_name)
            
            # Update combo
            self.sheet_combo['values'] = self.workbook.sheetnames
            self.sheet_var.set(sheet_name)
            
            # Show new sheet
            self.create_empty_sheet()
            self.status_label.config(text=f"Created new sheet: {sheet_name}")
            
            dialog.destroy()
        
        tk.Button(dialog, text="Create", command=create, width=10).pack(side="left", padx=40, pady=10)
        tk.Button(dialog, text="Cancel", command=dialog.destroy, width=10).pack(side="right", padx=40, pady=10)
        
        # Bind Enter key
        name_entry.bind("<Return>", lambda e: create())
        
    def create_ui(self):
        """Create user interface"""
        # Menubar
        menubar = tk.Menu(self.rootexcellite2)
        self.rootexcellite2.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New", command=self.create_empty_sheet)
        file_menu.add_command(label="New Sheet", command=self.create_new_sheet)
        file_menu.add_command(label="Open", command=lambda: self.open_excel_file())
        file_menu.add_command(label="Save", command=self.save_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.rootexcellite2.quit)
        
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Cut", command=self.cut_cell, accelerator="Ctrl+X")
        edit_menu.add_command(label="Copy", command=self.copy_cell, accelerator="Ctrl+C")
        edit_menu.add_command(label="Paste", command=self.paste_cell, accelerator="Ctrl+V")
        edit_menu.add_command(label="Delete", command=self.delete_cell, accelerator="Del")
        
        formulas_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Formulas", menu=formulas_menu)
        formulas_menu.add_command(label="Insert Formula", command=self.show_formula_dialog)
        formulas_menu.add_command(label="Recalculate All", command=self.recalculate_all)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
        help_menu.add_command(label="Formula Help", command=self.show_formula_help)
        
        # Keyboard shortcuts
        self.rootexcellite2.bind("<Control-c>", lambda e: self.copy_cell())
        self.rootexcellite2.bind("<Control-x>", lambda e: self.cut_cell())
        self.rootexcellite2.bind("<Control-v>", lambda e: self.paste_cell())
        self.rootexcellite2.bind("<Delete>", lambda e: self.delete_cell())
        
        # Toolbar
        toolbar_frame = tk.Frame(self.rootexcellite2, bg="#c0c0c0", relief="raised", bd=2, height=40)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        tk.Button(toolbar_frame, text="Open", bg="#c0c0c0", relief="raised", bd=2,
                 font=("MS Sans Serif", 8), command=lambda: self.open_excel_file()).pack(side="left", padx=5, pady=2)
        
        tk.Button(toolbar_frame, text="Save", bg="#c0c0c0", relief="raised", bd=2,
                 font=("MS Sans Serif", 8), command=self.save_excel_file).pack(side="left", padx=5, pady=2)
        
        tk.Label(toolbar_frame, text="Sheet:", bg="#c0c0c0", font=("MS Sans Serif", 8)).pack(side="left", padx=(20, 5))
        
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(toolbar_frame, textvariable=self.sheet_var, width=15, state="readonly")
        self.sheet_combo.pack(side="left", padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.show_sheet(self.sheet_var.get()))
        
        tk.Button(toolbar_frame, text="Formula", bg="#c0c0c0", relief="raised", bd=2,
                 font=("MS Sans Serif", 8), command=self.show_formula_dialog).pack(side="left", padx=5, pady=2)
        
        tk.Button(toolbar_frame, text="Calculate", bg="#c0c0c0", relief="raised", bd=2,
                 font=("MS Sans Serif", 8), command=self.recalculate_all).pack(side="left", padx=5, pady=2)
        
        # Formula bar
        formula_bar_frame = tk.Frame(self.rootexcellite2, bg="#c0c0c0", relief="sunken", bd=1, height=30)
        formula_bar_frame.pack(fill="x", padx=5, pady=2)
        formula_bar_frame.pack_propagate(False)
        
        self.cell_address_label = tk.Label(formula_bar_frame, text="", width=8, bg="white", 
                                           relief="sunken", bd=1, font=("MS Sans Serif", 9, "bold"))
        self.cell_address_label.pack(side="left", padx=2, pady=2)
        
        tk.Label(formula_bar_frame, text="fx", bg="#c0c0c0", font=("MS Sans Serif", 8)).pack(side="left", padx=2)
        
        self.formula_entry = tk.Entry(formula_bar_frame, font=("MS Sans Serif", 9), relief="sunken", bd=1)
        self.formula_entry.pack(side="left", fill="x", expand=True, padx=2, pady=2)
        self.formula_entry.bind("<Return>", self.on_formula_entry_return)
        self.formula_entry.bind("<FocusOut>", self.on_formula_entry_return)
        
        # Main content frame
        content_frame = tk.Frame(self.rootexcellite2, bg="white")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbars
        h_scrollbar = tk.Scrollbar(content_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        table_frame = tk.Frame(content_frame, bg="white")
        table_frame.pack(fill="both", expand=True)
        
        v_scrollbar = tk.Scrollbar(table_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        # Canvas for scrollable content
        self.canvas = tk.Canvas(table_frame, bg="white", 
                               xscrollcommand=h_scrollbar.set,
                               yscrollcommand=v_scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)

        h_scrollbar.config(command=self.canvas.xview)
        v_scrollbar.config(command=self.canvas.yview)

        # ÎNLOCUIEȘTE bind_all cu bind DOAR pe canvas:
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

        # Frame inside canvas for grid
        self.inner_frame = tk.Frame(self.canvas, bg="white")
        self.canvas.create_window((0, 0), window=self.inner_frame, anchor="nw")
        
        # Status bar
        status_frame = tk.Frame(self.rootexcellite2, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(status_frame, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        self.status_label.pack(side="left", padx=5)
        
        self.cell_info_label = tk.Label(status_frame, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        self.cell_info_label.pack(side="right", padx=5)
        
        # Create context menu
        self.create_context_menu()
        
        # Create empty sheet by default
        self.create_empty_sheet()
    
    def create_context_menu(self):
        """Create context menu for cells"""
        self.context_menu = tk.Menu(self.rootexcellite2, tearoff=0)
        self.context_menu.add_command(label="Cut", command=self.cut_cell, accelerator="Ctrl+X")
        self.context_menu.add_command(label="Copy", command=self.copy_cell, accelerator="Ctrl+C")
        self.context_menu.add_command(label="Paste", command=self.paste_cell, accelerator="Ctrl+V")
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Delete", command=self.delete_cell)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Insert Formula", command=self.show_formula_dialog)
    
    def show_context_menu(self, event, row, col):
        """Show context menu"""
        # Select the cell if not already selected
        if self.selected_cell != (row, col):
            self.select_cell(row, col)
        
        try:
            self.context_menu.tk_popup(event.x_rootexcellite2, event.y_rootexcellite2)
        finally:
            self.context_menu.grab_release()
    
    def delete_cell(self):
        """Delete content of selected cell(s)"""
        if not self.selected_range:
            return
        
        (start_row, start_col), (end_row, end_col) = self.selected_range
        min_row, max_row = min(start_row, end_row), max(start_row, end_row)
        min_col, max_col = min(start_col, end_col), max(start_col, end_col)
        
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if 0 <= r < len(self.cell_grid) and 0 <= c < len(self.cell_grid[0]):
                    self.cell_grid[r][c].delete(0, tk.END)
                    if (r, c) in self.cell_data:
                        del self.cell_data[(r, c)]
        
        self.status_label.config(text="Cell(s) deleted")
    
    def create_empty_sheet(self):
        """Create empty sheet"""
        self.clear_sheet()
        self.cell_data = {}
        
        self.header_labels = [None] * (self.MAX_COLS + 1)
        
        # Empty corner header
        header_label = tk.Label(self.inner_frame, text="", width=4, bg="#e0e0e0", 
                              relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
        header_label.grid(row=0, column=0, sticky="nsew")
        self.header_labels[0] = header_label
        
        # Column headers (A, B, C, etc.) with resize
        for col in range(self.MAX_COLS):
            col_letter = chr(65 + col)
            header = tk.Label(self.inner_frame, text=col_letter, width=self.column_widths[col+1], 
                            bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
            header.grid(row=0, column=col+1, sticky="nsew")
            self.header_labels[col+1] = header
            
            # Bind resize events
            header.bind("<Motion>", lambda e, c=col: self.on_header_motion(e, c))
            header.bind("<Button-1>", lambda e, c=col: self.on_header_click(e, c))
            header.bind("<B1-Motion>", lambda e, c=col: self.on_header_drag(e, c))
            header.bind("<ButtonRelease-1>", lambda e, c=col: self.on_header_release(e, c))
            header.bind("<Leave>", lambda e: self.reset_cursor())
        
        # Rows and cells
        row_widgets = []
        for row in range(self.MAX_ROWS):
            # Row header
            row_header = tk.Label(self.inner_frame, text=str(row + 1), width=4, height=1,
                                bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
            row_header.grid(row=row+1, column=0, sticky="nsew")
            
            # Cells
            col_widgets = []
            for col in range(self.MAX_COLS):
                cell = tk.Entry(self.inner_frame, width=self.column_widths[col+1], 
                              font=("MS Sans Serif", 9), relief="sunken", bd=1)
                cell.grid(row=row+1, column=col+1, sticky="nsew")
                cell.bind("<Button-1>", lambda e, r=row, c=col: self.select_cell(r, c, e))
                cell.bind("<Button-3>", lambda e, r=row, c=col: self.show_context_menu(e, r, c))
                cell.bind("<Shift-Button-1>", lambda e, r=row, c=col: self.extend_selection(r, c))
                cell.bind("<KeyRelease>", lambda e, r=row, c=col: self.on_cell_change(r, c, e))
                col_widgets.append(cell)
            row_widgets.append(col_widgets)
        
        self.cell_grid = row_widgets
        
        # Update canvas scroll region
        self.inner_frame.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        
        self.status_label.config(text="New spreadsheet created")
    
    def on_header_motion(self, event, col):
        """Detect when mouse is at header edge"""
        widget = event.widget
        widget_width = widget.winfo_width()
        
        # Check if we're near the right edge
        if widget_width - event.x <= self.resize_grip_width:
            widget.config(cursor="sb_h_double_arrow")
        else:
            widget.config(cursor="")
    
    def reset_cursor(self):
        """Reset cursor when leaving header"""
        # Cursor will be reset automatically when mouse leaves widget
        pass
    
    def on_header_click(self, event, col):
        """Start column resize if at right edge"""
        widget = event.widget
        widget_width = widget.winfo_width()
        
        # Check if click is near the right edge
        if widget_width - event.x <= self.resize_grip_width:
            self.resize_column = col
            self.resize_start_x = event.x_rootexcellite2
            self.resize_start_width = self.column_widths[col + 1]
    
    def on_header_drag(self, event, col):
        """Resize column"""
        if self.resize_column == col and self.resize_start_x is not None:
            delta = event.x_rootexcellite2 - self.resize_start_x
            new_width = max(5, self.resize_start_width + delta // 7)  # Minimum width 5
            self.column_widths[col + 1] = new_width
            
            # Update header width
            self.header_labels[col + 1].config(width=new_width)
            
            # Update all cells in this column
            for row in range(len(self.cell_grid)):
                if col < len(self.cell_grid[row]):
                    self.cell_grid[row][col].config(width=new_width)
            
            # Update canvas scroll region
            self.inner_frame.update_idletasks()
            self.canvas.config(scrollregion=self.canvas.bbox("all"))
    
    def on_header_release(self, event, col):
        """End column resize"""
        self.resize_column = None
        self.resize_start_x = None
        self.resize_start_width = None
    
    def clear_sheet(self):
        """Clear current sheet"""
        for widget in self.inner_frame.winfo_children():
            widget.destroy()
        self.cell_grid = []
        self.header_labels = []
    
    def select_cell(self, row, col, event=None):
        """Select a cell"""
        # Clear previous selection highlight
        self.clear_selection_highlight()
        
        self.selected_cell = (row, col)
        self.selected_range = [(row, col), (row, col)]
        
        # Highlight selected cell
        if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
            self.cell_grid[row][col].config(bg="#b0d4f1")
        
        col_letter = chr(65 + col)
        cell_address = f"{col_letter}{row + 1}"
        self.cell_address_label.config(text=cell_address)
        self.cell_info_label.config(text=f"Cell: {cell_address}")
        
        # Update formula bar
        if (row, col) in self.cell_data and 'formula' in self.cell_data[(row, col)]:
            self.formula_entry.delete(0, tk.END)
            self.formula_entry.insert(0, self.cell_data[(row, col)]['formula'])
        else:
            self.formula_entry.delete(0, tk.END)
            if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
                self.formula_entry.insert(0, self.cell_grid[row][col].get())
        
        # Set focus
        if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
            self.cell_grid[row][col].focus_set()
    
    def extend_selection(self, row, col):
        """Extend selection with Shift+Click"""
        if not self.selected_cell:
            self.select_cell(row, col)
            return
        
        self.clear_selection_highlight()
        
        start_row, start_col = self.selected_cell
        self.selected_range = [(start_row, start_col), (row, col)]
        
        # Highlight range
        min_row = min(start_row, row)
        max_row = max(start_row, row)
        min_col = min(start_col, col)
        max_col = max(start_col, col)
        
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if 0 <= r < len(self.cell_grid) and 0 <= c < len(self.cell_grid[0]):
                    self.cell_grid[r][c].config(bg="#b0d4f1")
        
        self.cell_info_label.config(text=f"Selection: {chr(65+min_col)}{min_row+1}:{chr(65+max_col)}{max_row+1}")
    
    def clear_selection_highlight(self):
        """Clear selection highlighting"""
        for row in range(len(self.cell_grid)):
            for col in range(len(self.cell_grid[0])):
                self.cell_grid[row][col].config(bg="white")
    
    def on_cell_change(self, row, col, event):
        """Handle cell content change"""
        if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
            value = self.cell_grid[row][col].get()
            
            if value.startswith('='):
                # It's a formula
                self.cell_data[(row, col)] = {'formula': value, 'value': value}
                self.evaluate_cell_formula(row, col)
            else:
                # Simple value
                if (row, col) in self.cell_data:
                    self.cell_data[(row, col)]['value'] = value
                    if 'formula' in self.cell_data[(row, col)]:
                        del self.cell_data[(row, col)]['formula']
                else:
                    self.cell_data[(row, col)] = {'value': value}
    
    def on_formula_entry_return(self, event):
        """Handle formula bar return/focus out"""
        if self.selected_cell:
            row, col = self.selected_cell
            formula_text = self.formula_entry.get()
            
            if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
                self.cell_grid[row][col].delete(0, tk.END)
                self.cell_grid[row][col].insert(0, formula_text)
                
                if formula_text.startswith('='):
                    self.cell_data[(row, col)] = {'formula': formula_text, 'value': formula_text}
                    self.evaluate_cell_formula(row, col)
                else:
                    self.cell_data[(row, col)] = {'value': formula_text}
    
    def evaluate_cell_formula(self, row, col):
        """Evaluate formula in a cell"""
        if (row, col) not in self.cell_data or 'formula' not in self.cell_data[(row, col)]:
            return
        
        formula = self.cell_data[(row, col)]['formula']
        
        try:
            result = self.evaluate_formula(formula, row, col)
            self.cell_grid[row][col].delete(0, tk.END)
            
            # Format result for display
            if isinstance(result, float):
                # Display with appropriate precision
                if result == int(result):
                    self.cell_grid[row][col].insert(0, str(int(result)))
                else:
                    self.cell_grid[row][col].insert(0, f"{result:.10g}")
            else:
                self.cell_grid[row][col].insert(0, str(result))
            
            self.cell_data[(row, col)]['value'] = result
        except Exception as e:
            error_msg = "#ERROR"
            if "division by zero" in str(e).lower() or "zerodivision" in str(e).lower():
                error_msg = "#DIV/0!"
            elif "name" in str(e).lower():
                error_msg = "#NAME?"
            elif "value" in str(e).lower():
                error_msg = "#VALUE!"
            
            self.cell_grid[row][col].delete(0, tk.END)
            self.cell_grid[row][col].insert(0, error_msg)
            self.cell_data[(row, col)]['value'] = error_msg
            print(f"Formula error in {chr(65+col)}{row+1}: {str(e)}")
    
    def parse_range(self, range_str):
        """Parse a range like A1:A10 and return list of values (excludes empty cells)"""
        if ':' not in range_str:
            val = self.get_cell_value_by_ref(range_str)
            # Only include non-empty values
            if val not in [None, '']:
                return [val]
            return []
        
        parts = range_str.split(':')
        if len(parts) != 2:
            return []
        
        start_ref, end_ref = parts
        
        # Parse start cell
        match_start = re.match(r'([A-Z]+)(\d+)', start_ref)
        match_end = re.match(r'([A-Z]+)(\d+)', end_ref)
        
        if not match_start or not match_end:
            return []
        
        start_col_letter, start_row_num = match_start.groups()
        end_col_letter, end_row_num = match_end.groups()
        
        start_col = ord(start_col_letter) - 65
        end_col = ord(end_col_letter) - 65
        start_row = int(start_row_num) - 1
        end_row = int(end_row_num) - 1
        
        values = []
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                val = self.get_cell_value_by_ref(f"{chr(65+c)}{r+1}")
                # Only include non-empty and non-error values in ranges
                if val is not None and not (isinstance(val, str) and val.startswith('#')):
                    values.append(val)

        return values
    
    def evaluate_formula(self, formula, row=None, col=None):
        """Evaluate an Excel formula"""
        if not formula.startswith('='):
            return formula
        
        def safe_divide(numerator, denominator, error_value=0):
            """Safe division that returns error_value instead of crashing"""
            try:
                if denominator == 0 or denominator == 0.0:
                    return error_value
                return numerator / denominator
            except:
                return error_value

        def iferror(formula_func, error_value=0):
            """IFERROR implementation - executes formula and returns error_value if it fails"""
            try:
                result = formula_func()
                # Check if result is an error string
                if isinstance(result, str) and result.startswith('#'):
                    return error_value
                return result
            except:
                return error_value
                
        formula_original = formula[1:]  # Remove '='
        formula = formula_original.upper()
        
        # Replace range references in functions - IMPORTANT: must keep as list notation
        def replace_range(match):
            range_str = match.group(0)
            values = self.parse_range(range_str)
            # Return as list literal that Python can evaluate
            if not values:
                return '[]'
            return '[' + ','.join(str(v) if isinstance(v, (int, float)) else f'"{v}"' for v in values) + ']'

        # Replace ranges like A1:A10 with [val1, val2, ...]
        formula = re.sub(r'[A-Z]+\d+:[A-Z]+\d+', replace_range, formula)

        # Replace single cell references (but not those already in ranges)
        def replace_cell_ref(match):
            cell_ref = match.group(0)
            value = self.get_cell_value_by_ref(cell_ref)
            # Return None for empty cells (Excel ignores them in math)
            if value is None or value == '':
                return '0'
            # Ignore error values
            if isinstance(value, str) and value.startswith('#'):
                return '0'
            # If it's a string, put it in quotes
            if isinstance(value, str):
                return f'"{value}"'
            return str(value)

        # Only replace standalone cell references (not part of ranges)
        formula = re.sub(r'(?<![A-Z0-9:])[A-Z]+\d+(?!:)', replace_cell_ref, formula)
        
        # Define helper functions
        def excel_sum(*args):
            """Custom SUM that handles both lists and individual values, ignores empty and errors"""
            result = 0
            for arg in args:
                if arg is None:
                    continue
                if isinstance(arg, (list, tuple)):
                    for x in arg:
                        if x is not None and x != '':
                            if isinstance(x, (int, float)):
                                result += x
                elif isinstance(arg, (int, float)):
                    result += arg
                elif isinstance(arg, str) and arg.startswith('#'):
                    continue  # Ignore errors
                elif arg != '' and arg is not None:
                    try:
                        result += float(arg)
                    except:
                        pass
            return result

        def excel_average(*args):
            """Custom AVERAGE function, ignores empty strings and errors"""
            flat_args = []
            for arg in args:
                if arg is None:
                    continue
                if isinstance(arg, (list, tuple)):
                    for x in arg:
                        if x is not None and x != '':
                            flat_args.append(x)
                else:
                    flat_args.append(arg)
            numeric_args = []
            for x in flat_args:
                if x is None or (isinstance(x, str) and x.startswith('#')):
                    continue  # Ignore None and errors
                if x != '' and x is not None:
                    try:
                        numeric_args.append(float(x))
                    except:
                        pass
            return sum(numeric_args) / len(numeric_args) if numeric_args else 0
        
        def excel_min(*args):
            """Custom MIN function"""
            flat_args = []
            for arg in args:
                if isinstance(arg, (list, tuple)):
                    flat_args.extend(arg)
                else:
                    flat_args.append(arg)
            numeric_args = []
            for x in flat_args:
                try:
                    numeric_args.append(float(x))
                except:
                    pass
            return min(numeric_args) if numeric_args else 0
        
        def excel_max(*args):
            """Custom MAX function"""
            flat_args = []
            for arg in args:
                if isinstance(arg, (list, tuple)):
                    flat_args.extend(arg)
                else:
                    flat_args.append(arg)
            numeric_args = []
            for x in flat_args:
                try:
                    numeric_args.append(float(x))
                except:
                    pass
            return max(numeric_args) if numeric_args else 0
        
        def excel_count(*args):
            """Custom COUNT function - counts non-empty numeric values"""
            flat_args = []
            for arg in args:
                if isinstance(arg, (list, tuple)):
                    flat_args.extend(arg)
                else:
                    flat_args.append(arg)
            count = 0
            for x in flat_args:
                if x != '' and x is not None:  # Count non-empty values
                    try:
                        float(x)  # Check if it's numeric
                        count += 1
                    except:
                        pass
            return count
        
        def sumif(range_vals, criteria):
            """SUMIF function"""
            if not isinstance(range_vals, (list, tuple)):
                range_vals = [range_vals]
            total = 0
            for val in range_vals:
                try:
                    if eval(f"{val}{criteria}"):
                        total += float(val)
                except:
                    pass
            return total
        
        def countif(range_vals, criteria):
            """COUNTIF function"""
            if not isinstance(range_vals, (list, tuple)):
                range_vals = [range_vals]
            count = 0
            for val in range_vals:
                try:
                    if eval(f"{val}{criteria}"):
                        count += 1
                except:
                    pass
            return count
        
        def if_func(condition, true_val, false_val):
            """IF function"""
            return true_val if condition else false_val
        
        def excel_round(number, num_digits=0):
            """ROUND function"""
            try:
                return round(float(number), int(num_digits))
            except:
                return 0
        
        def excel_roundup(number, num_digits=0):
            """ROUNDUP function"""
            try:
                import math
                multiplier = 10 ** int(num_digits)
                return math.ceil(float(number) * multiplier) / multiplier
            except:
                return 0
        
        def excel_rounddown(number, num_digits=0):
            """ROUNDDOWN function"""
            try:
                import math
                multiplier = 10 ** int(num_digits)
                return math.floor(float(number) * multiplier) / multiplier
            except:
                return 0
        
        def excel_mod(number, divisor):
            """MOD function"""
            try:
                return float(number) % float(divisor)
            except:
                return 0
        
        def excel_floor(number):
            """FLOOR function"""
            try:
                import math
                return math.floor(float(number))
            except:
                return 0
        
        def excel_ceiling(number):
            """CEILING function"""
            try:
                import math
                return math.ceil(float(number))
            except:
                return 0
        
        def excel_len(text):
            """LEN function"""
            return len(str(text))
        
        def excel_upper(text):
            """UPPER function"""
            return str(text).upper()
        
        def excel_lower(text):
            """LOWER function"""
            return str(text).lower()
        
        def excel_trim(text):
            """TRIM function"""
            return str(text).strip()
        
        def excel_left(text, num_chars=1):
            """LEFT function"""
            return str(text)[:int(num_chars)]
        
        def excel_right(text, num_chars=1):
            """RIGHT function"""
            return str(text)[-int(num_chars):]
        
        def excel_mid(text, start_num, num_chars):
            """MID function"""
            start = int(start_num) - 1  # Excel uses 1-based indexing
            return str(text)[start:start + int(num_chars)]
        
        def excel_concatenate(*args):
            """CONCATENATE function"""
            return ''.join(str(arg) for arg in args)
        
        def excel_log(number, base=10):
            """LOG function"""
            try:
                import math
                return math.log(float(number), float(base))
            except:
                return 0
        
        def excel_ln(number):
            """LN function"""
            try:
                import math
                return math.log(float(number))
            except:
                return 0
        
        def excel_exp(number):
            """EXP function"""
            try:
                import math
                return math.exp(float(number))
            except:
                return 0
        
        def excel_sin(number):
            """SIN function"""
            try:
                import math
                return math.sin(float(number))
            except:
                return 0
        
        def excel_cos(number):
            """COS function"""
            try:
                import math
                return math.cos(float(number))
            except:
                return 0
        
        def excel_tan(number):
            """TAN function"""
            try:
                import math
                return math.tan(float(number))
            except:
                return 0
        
        def excel_radians(degrees):
            """RADIANS function"""
            try:
                import math
                return math.radians(float(degrees))
            except:
                return 0
        
        def excel_degrees(radians):
            """DEGREES function"""
            try:
                import math
                return math.degrees(float(radians))
            except:
                return 0
        
        def excel_rand():
            """RAND function"""
            import random
            return random.random()
        
        def excel_randbetween(bottom, top):
            """RANDBETWEEN function"""
            import random
            return random.randint(int(bottom), int(top))
        
        def excel_product(*args):
            """PRODUCT function"""
            result = 1
            for arg in args:
                if isinstance(arg, (list, tuple)):
                    result *= excel_product(*arg)
                elif isinstance(arg, (int, float)):
                    result *= arg
                else:
                    try:
                        result *= float(arg)
                    except:
                        pass
            return result
        
        def excel_not(logical):
            """NOT function"""
            return not logical
        
        def excel_and(*args):
            """AND function"""
            return all(args)
        
        def excel_or(*args):
            """OR function"""
            return any(args)
        
        # Replace Excel functions with custom equivalents
        formula = formula.replace('SUM(', 'excel_sum(')
        formula = formula.replace('AVERAGE(', 'excel_average(')
        formula = formula.replace('AVG(', 'excel_average(')
        formula = formula.replace('MIN(', 'excel_min(')
        formula = formula.replace('MAX(', 'excel_max(')
        formula = formula.replace('COUNT(', 'excel_count(')
        formula = formula.replace('SUMIF(', 'sumif(')
        formula = formula.replace('COUNTIF(', 'countif(')
        formula = formula.replace('IF(', 'if_func(')
        formula = formula.replace('ROUND(', 'excel_round(')
        formula = formula.replace('ROUNDUP(', 'excel_roundup(')
        formula = formula.replace('ROUNDDOWN(', 'excel_rounddown(')
        formula = formula.replace('MOD(', 'excel_mod(')
        formula = formula.replace('FLOOR(', 'excel_floor(')
        formula = formula.replace('CEILING(', 'excel_ceiling(')
        formula = formula.replace('LEN(', 'excel_len(')
        formula = formula.replace('UPPER(', 'excel_upper(')
        formula = formula.replace('LOWER(', 'excel_lower(')
        formula = formula.replace('TRIM(', 'excel_trim(')
        formula = formula.replace('LEFT(', 'excel_left(')
        formula = formula.replace('RIGHT(', 'excel_right(')
        formula = formula.replace('MID(', 'excel_mid(')
        formula = formula.replace('CONCATENATE(', 'excel_concatenate(')
        formula = formula.replace('LOG(', 'excel_log(')
        formula = formula.replace('LN(', 'excel_ln(')
        formula = formula.replace('EXP(', 'excel_exp(')
        formula = formula.replace('SIN(', 'excel_sin(')
        formula = formula.replace('COS(', 'excel_cos(')
        formula = formula.replace('TAN(', 'excel_tan(')
        formula = formula.replace('RADIANS(', 'excel_radians(')
        formula = formula.replace('DEGREES(', 'excel_degrees(')
        formula = formula.replace('RAND(', 'excel_rand(')
        formula = formula.replace('RANDBETWEEN(', 'excel_randbetween(')
        formula = formula.replace('PRODUCT(', 'excel_product(')
        formula = formula.replace('NOT(', 'excel_not(')
        formula = formula.replace('AND(', 'excel_and(')
        formula = formula.replace('OR(', 'excel_or(')
        formula = formula.replace('SQRT(', 'sqrt(')
        formula = formula.replace('POWER(', 'pow(')
        formula = formula.replace('ABS(', 'abs(')
        formula = formula.replace('IFERROR(', 'iferror(lambda: ')
        
        # Evaluate formula
        try:
            import math
            result = eval(formula, {"__builtins__": {}}, {
                'excel_sum': excel_sum,
                'excel_average': excel_average,
                'excel_min': excel_min,
                'excel_max': excel_max,
                'excel_count': excel_count,
                'excel_round': excel_round,
                'excel_roundup': excel_roundup,
                'excel_rounddown': excel_rounddown,
                'excel_mod': excel_mod,
                'excel_floor': excel_floor,
                'excel_ceiling': excel_ceiling,
                'excel_len': excel_len,
                'excel_upper': excel_upper,
                'excel_lower': excel_lower,
                'excel_trim': excel_trim,
                'excel_left': excel_left,
                'excel_right': excel_right,
                'excel_mid': excel_mid,
                'excel_concatenate': excel_concatenate,
                'excel_log': excel_log,
                'excel_ln': excel_ln,
                'excel_exp': excel_exp,
                'excel_sin': excel_sin,
                'excel_cos': excel_cos,
                'excel_tan': excel_tan,
                'excel_radians': excel_radians,
                'excel_degrees': excel_degrees,
                'excel_rand': excel_rand,
                'excel_randbetween': excel_randbetween,
                'excel_product': excel_product,
                'excel_not': excel_not,
                'excel_and': excel_and,
                'excel_or': excel_or,
                'abs': abs,
                'round': round,
                'int': int,
                'float': float,
                'str': str,
                'len': len,
                'sumif': sumif,
                'countif': countif,
                'if_func': if_func,
                'sqrt': math.sqrt,
                'pow': pow,
                'pi': math.pi,
                'e': math.e
            })
            
            if isinstance(result, float):
                if math.isnan(result) or math.isinf(result):
                    return "#DIV/0!"
            
            # Handle division by zero specifically - return 0 or blank instead of error
            return result
            
        except ZeroDivisionError:
            # Return Excel-style DIV/0 error
            return "#DIV/0!"
            
        except Exception as e:
            raise Exception(f"Formula error: {str(e)}")
    def get_cell_value_by_ref(self, cell_ref):
        """Get cell value from its reference"""
        match = re.match(r'([A-Z]+)(\d+)', cell_ref.upper())
        if not match:
            return None
        
        col_letter, row_num = match.groups()
        col = ord(col_letter) - 65
        row = int(row_num) - 1
        
        if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
            if (row, col) in self.cell_data:
                value = self.cell_data[(row, col)].get('value', '')
                # Ignore error values
                if isinstance(value, str) and value.startswith('#'):
                    return None
                # Return the actual value
                if isinstance(value, (int, float)):
                    return value
                if isinstance(value, str) and value:
                    try:
                        return float(value) if '.' in str(value) else int(value)
                    except:
                        return value
                return None  # Return None for empty cells
            else:
                value = self.cell_grid[row][col].get()
                if not value:
                    return None  # Return None for empty cells
                try:
                    return float(value) if '.' in value else int(value)
                except:
                    return value if value else None
        return None  # Return None instead of empty string
    
    def recalculate_all(self):
        """Recalculate all formulas"""
        for (row, col), data in self.cell_data.items():
            if 'formula' in data:
                self.evaluate_cell_formula(row, col)
        self.status_label.config(text="All formulas recalculated")
    
    def show_formula_dialog(self):
        """Show dialog for inserting formula"""
        if not self.selected_cell:
            messagebox.showinfo("Info", "Please select a cell first")
            return
        
        dialog = tk.Toplevel(self.rootexcellite2)
        dialog.title("Insert Formula")
        dialog.geometry("450x600")
        dialog.transient(self.rootexcellite2)
        dialog.grab_set()
        
        tk.Label(dialog, text="Enter formula (start with =):", font=("MS Sans Serif", 9)).pack(pady=10)
        
        formula_text = tk.Text(dialog, height=4, width=50, font=("MS Sans Serif", 9))
        formula_text.pack(pady=5, padx=10)
        
        # Formula examples
        examples_frame = tk.LabelFrame(dialog, text="Formula Examples", font=("MS Sans Serif", 8))
        examples_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        examples = [
            "=A1+B1                    (Addition)",
            "=SUM(A1:A10)              (Sum range)",
            "=AVERAGE(B1:B5)           (Average)",
            "=MAX(C1:C10)              (Maximum value)",
            "=MIN(C1:C10)              (Minimum value)",
            "=ROUND(A1/B1, 2)          (Round to 2 decimals)",
            "=IF(A1>10, \"Yes\", \"No\")  (Conditional)",
            "=CONCATENATE(A1,\" \",B1) (Join text)",
            "=UPPER(A1)                (Uppercase)",
            "=LEN(A1)                  (Text length)",
            "=SQRT(A1)                 (Square rootexcellite2)",
            "=RAND()                   (Random 0-1)",
            "=SUMIF(A1:A10, \">5\")    (Conditional sum)",
            "=COUNTIF(A1:A10, \"<>0\") (Count non-zero)"
        ]
        
        for example in examples:
            tk.Label(examples_frame, text=example, font=("Courier", 8), fg="blue", anchor="w").pack(anchor="w", padx=5, pady=1)
        
        def insert_formula():
            formula = formula_text.get("1.0", tk.END).strip()
            if formula:
                row, col = self.selected_cell
                self.cell_grid[row][col].delete(0, tk.END)
                self.cell_grid[row][col].insert(0, formula)
                self.cell_data[(row, col)] = {'formula': formula, 'value': formula}
                self.evaluate_cell_formula(row, col)
                dialog.destroy()
        
        tk.Button(dialog, text="Insert", command=insert_formula, width=10).pack(side="left", padx=50, pady=10)
        tk.Button(dialog, text="Cancel", command=dialog.destroy, width=10).pack(side="right", padx=50, pady=10)
    
    def show_formula_help(self):
        """Show formula help in scrollable window"""
        dialog = tk.Toplevel(self.rootexcellite2)
        dialog.title("Formula Help")
        dialog.geometry("700x600")
        dialog.transient(self.rootexcellite2)
        
        # Make resizable
        dialog.resizable(True, True)
        
        # Create frame with scrollbar
        main_frame = tk.Frame(dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(main_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Text widget
        text_widget = tk.Text(main_frame, wrap="word", yscrollcommand=scrollbar.set,
                             font=("Courier New", 10), padx=10, pady=10)
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)
        
        help_text = """Excel Lite - Formula Help

Supported Functions:

MATHEMATICAL:
- SUM(range) - Sum of values
- AVERAGE(range) - Average of values  
- MIN(range) - Minimum value
- MAX(range) - Maximum value
- COUNT(range) - Count of values
- PRODUCT(range) - Product of values
- ROUND(number, digits) - Round to digits
- ROUNDUP(number, digits) - Round up
- ROUNDDOWN(number, digits) - Round down
- MOD(number, divisor) - Modulo
- FLOOR(number) - Round down to integer
- CEILING(number) - Round up to integer
- SQRT(number) - Square rootexcellite2
- POWER(number, power) - Power function
- ABS(number) - Absolute value

CONDITIONAL:
- SUMIF(range, criteria) - Conditional sum
- COUNTIF(range, criteria) - Conditional count
- IF(condition, true_val, false_val) - Conditional

TEXT:
- LEN(text) - Length of text
- UPPER(text) - Convert to uppercase
- LOWER(text) - Convert to lowercase
- TRIM(text) - Remove spaces
- LEFT(text, n) - Left n characters
- RIGHT(text, n) - Right n characters
- MID(text, start, n) - Middle n characters
- CONCATENATE(text1, text2, ...) - Join text

TRIGONOMETRIC:
- SIN(radians) - Sine
- COS(radians) - Cosine
- TAN(radians) - Tangent
- RADIANS(degrees) - Convert to radians
- DEGREES(radians) - Convert to degrees

LOGARITHMIC:
- LOG(number, base) - Logarithm
- LN(number) - Natural logarithm
- EXP(number) - e raised to power

LOGICAL:
- AND(logical1, logical2, ...) - Logical AND
- OR(logical1, logical2, ...) - Logical OR
- NOT(logical) - Logical NOT

RANDOM:
- RAND() - Random between 0 and 1
- RANDBETWEEN(min, max) - Random integer

Operators:
- + - * / (arithmetic)
- > < >= <= = <> (comparisons)
- () (parentheses for grouping)

Examples:
- =A1+B1
- =SUM(A1:A10)
- =AVERAGE(B1:B5)
- =IF(A1>100, "High", "Low")
- =ROUND(A1/B1, 2)
- =CONCATENATE(A1, " ", B1)
- =UPPER(A1)
- =SQRT(POWER(A1,2)+POWER(B1,2))
- =SUMIF(A1:A10, ">5")
- =COUNTIF(A1:A10, "<>0")
    """
        
        text_widget.insert("1.0", help_text)
        text_widget.config(state="disabled")  # Make read-only
        
        # Close button
        tk.Button(dialog, text="Close", command=dialog.destroy, width=15).pack(pady=10)
    
    def cut_cell(self):
        """Cut selected cell or range"""
        if not self.selected_range:
            return
        
        (start_row, start_col), (end_row, end_col) = self.selected_range
        min_row, max_row = min(start_row, end_row), max(start_row, end_row)
        min_col, max_col = min(start_col, end_col), max(start_col, end_col)
        
        self.clipboard_data = []
        for r in range(min_row, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                if 0 <= r < len(self.cell_grid) and 0 <= c < len(self.cell_grid[0]):
                    value = self.cell_grid[r][c].get()
                    row_data.append(value)
                    self.cell_grid[r][c].delete(0, tk.END)
                    if (r, c) in self.cell_data:
                        del self.cell_data[(r, c)]
            self.clipboard_data.append(row_data)
        
        self.status_label.config(text=f"Cut {len(self.clipboard_data)}x{len(self.clipboard_data[0])} cells")
    
    def copy_cell(self):
        """Copy selected cell or range"""
        if not self.selected_range:
            return
        
        (start_row, start_col), (end_row, end_col) = self.selected_range
        min_row, max_row = min(start_row, end_row), max(start_row, end_row)
        min_col, max_col = min(start_col, end_col), max(start_col, end_col)
        
        self.clipboard_data = []
        for r in range(min_row, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                if 0 <= r < len(self.cell_grid) and 0 <= c < len(self.cell_grid[0]):
                    value = self.cell_grid[r][c].get()
                    row_data.append(value)
            self.clipboard_data.append(row_data)
        
        self.status_label.config(text=f"Copied {len(self.clipboard_data)}x{len(self.clipboard_data[0])} cells")
    
    def paste_cell(self):
        """Paste to selected cell"""
        if not self.selected_cell or not self.clipboard_data:
            return
        
        start_row, start_col = self.selected_cell
        
        for i, row_data in enumerate(self.clipboard_data):
            for j, value in enumerate(row_data):
                row = start_row + i
                col = start_col + j
                if 0 <= row < len(self.cell_grid) and 0 <= col < len(self.cell_grid[0]):
                    self.cell_grid[row][col].delete(0, tk.END)
                    self.cell_grid[row][col].insert(0, value)
                    
                    if value.startswith('='):
                        self.cell_data[(row, col)] = {'formula': value, 'value': value}
                        self.evaluate_cell_formula(row, col)
                    else:
                        self.cell_data[(row, col)] = {'value': value}
        
        self.status_label.config(text=f"Pasted {len(self.clipboard_data)}x{len(self.clipboard_data[0])} cells")
    
    def open_excel_file(self, file_path=None):
        """Open an Excel file"""
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Open Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
        
        if not file_path:
            return
        
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.current_file = file_path
            
            # Populate sheet combo
            sheet_names = self.workbook.sheetnames
            self.sheet_combo['values'] = sheet_names
            
            if sheet_names:
                self.sheet_var.set(sheet_names[0])
                self.show_sheet(sheet_names[0])
            
            self.status_label.config(text=f"Opened: {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{str(e)}")
    
    def show_sheet(self, sheet_name):
        """Display a specific sheet"""
        if not self.workbook:
            return
        
        try:
            self.active_sheet = self.workbook[sheet_name]
            
            # Determină numărul real de rânduri din sheet
            actual_max_row = self.active_sheet.max_row
            
            # Actualizează MAX_ROWS ÎNAINTE de a crea grid-ul
            if actual_max_row > self.MAX_ROWS:
                self.MAX_ROWS = actual_max_row
            
            self.clear_sheet()
            self.cell_data = {}
            
            # Recreate grid cu numărul corect de rânduri (acum MAX_ROWS este actualizat)
            self.create_empty_sheet()
            
            # Load data from sheet
            for row_idx, row in enumerate(self.active_sheet.iter_rows(max_row=actual_max_row, max_col=self.MAX_COLS)):
                for col_idx, cell in enumerate(row):
                    if cell.value is not None:
                        # ADAUGĂ verificare suplimentară
                        if row_idx < len(self.cell_grid) and col_idx < len(self.cell_grid[0]):
                            value = str(cell.value)
                            self.cell_grid[row_idx][col_idx].delete(0, tk.END)
                            self.cell_grid[row_idx][col_idx].insert(0, value)
                            
                            if value.startswith('='):
                                self.cell_data[(row_idx, col_idx)] = {'formula': value, 'value': value}
                                self.evaluate_cell_formula(row_idx, col_idx)
                            else:
                                self.cell_data[(row_idx, col_idx)] = {'value': value}
            
            # Recalculează toate formulele după încărcare
            self.recalculate_all()
            
            self.status_label.config(text=f"Showing sheet: {sheet_name}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load sheet:\n{str(e)}")
    
    def save_excel_file(self):
        """Save current sheet to Excel file"""
        file_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Save data
            for row_idx in range(len(self.cell_grid)):
                for col_idx in range(len(self.cell_grid[0])):
                    value = self.cell_grid[row_idx][col_idx].get()
                    if value:
                        # Check if it's a formula
                        if (row_idx, col_idx) in self.cell_data and 'formula' in self.cell_data[(row_idx, col_idx)]:
                            ws.cell(row=row_idx+1, column=col_idx+1, value=self.cell_data[(row_idx, col_idx)]['formula'])
                        else:
                            # Try to convert to number if possible
                            try:
                                numeric_value = float(value)
                                ws.cell(row=row_idx+1, column=col_idx+1, value=numeric_value)
                            except:
                                ws.cell(row=row_idx+1, column=col_idx+1, value=value)
            
            wb.save(file_path)
            self.current_file = file_path
            self.status_label.config(text=f"Saved: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "File saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file:\n{str(e)}")
    
    def show_about(self):
        """Show about dialog in scrollable window"""
        dialog = tk.Toplevel(self.rootexcellite2)
        dialog.title("About Excel Lite")
        dialog.geometry("500x400")
        dialog.transient(self.rootexcellite2)
        
        # Make resizable
        dialog.resizable(True, True)
        
        # Create frame with scrollbar
        main_frame = tk.Frame(dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(main_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Text widget
        text_widget = tk.Text(main_frame, wrap="word", yscrollcommand=scrollbar.set,
                             font=("Arial", 11), padx=20, pady=20)
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)
        
        about_text = """Excel Lite v1.0

A lightweight spreadsheet application.

Features:
- Create and edit spreadsheets
- Support for Excel formulas
- Open and save Excel files (.xlsx)
- Copy, cut, paste operations
- Cell range selection
- Column resizing
- Multiple sheets support
- Automatic formula calculation
- Dynamic row sizing

Supported Formulas:
- Mathematical: SUM, AVERAGE, MIN, MAX, COUNT, etc.
- Conditional: IF, SUMIF, COUNTIF
- Text: UPPER, LOWER, LEN, CONCATENATE, etc.
- Trigonometric: SIN, COS, TAN
- Logical: AND, OR, NOT
- And many more...

Keyboard Shortcuts:
- Ctrl+C - Copy
- Ctrl+X - Cut
- Ctrl+V - Paste
- Delete - Delete cell content


© 2025 Excel Lite
    """
        
        text_widget.insert("1.0", about_text)
        text_widget.config(state="disabled")  # Make read-only
        
        # Close button
        tk.Button(dialog, text="Close", command=dialog.destroy, width=15).pack(pady=10)


def main():
    rootexcellite2 = tk.Tk()
    app = ExcelLite(rootexcellite2)
    rootexcellite2.mainloop()


if __name__ == "__main__":
    main()